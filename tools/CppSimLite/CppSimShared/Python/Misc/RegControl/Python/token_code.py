import ctypes as ct
import numpy as np
import sys
import os
import platform
import openpyxl

reg_control_home = os.getenv("REG_CONTROL_HOME")

class TokenMap(object):
    def __init__(self, filename=None, verbose_messages=None):
       self.err_msg = ct.create_string_buffer(10000)
       self.out_string = ct.create_string_buffer(1000)
       if sys.platform == 'darwin':
          token_code_lib_file = '%s/macos/token_code_lib.so' % reg_control_home
       elif sys.platform == 'win32':
          if sys.maxsize == 2147483647:
             token_code_lib_file = '%s/win32/token_code_lib.dll' % reg_control_home
          else:
             token_code_lib_file = '%s/win64/token_code_lib.dll' % reg_control_home
       else:
          if arch_val == '64bit':
             token_code_lib_file = '%s/glnxa64/token_code_lib.so' % reg_control_home
          else:
             token_code_lib_file = '%s/glnx86/token_code_lib.so' % reg_control_home

       self.token_code_lib = ct.CDLL(token_code_lib_file)
       self.token_code_lib.initialize.argtypes = [ct.c_char_p]
       self.token_code_lib.add_token.argtypes = [ct.c_char_p, ct.c_char_p, ct.c_char_p, ct.c_char_p, ct.c_int, ct.c_char_p, ct.c_char_p, ct.c_char_p]
       self.token_code_lib.add_token_with_test_pad_info.argtypes = [ct.c_char_p, ct.c_char_p, ct.c_char_p, ct.c_char_p, ct.c_int, ct.c_char_p, ct.c_char_p, ct.c_char_p, ct.c_char_p, ct.c_char_p, ct.c_char_p, ct.c_char_p]
       self.token_code_lib.add_formula.argtypes = [ct.c_char_p, ct.c_char_p, ct.c_char_p, ct.c_int, ct.c_int, ct.c_char_p, ct.c_char_p, ct.c_char_p]
       self.token_code_lib.add_inv_formula.argtypes = [ct.c_char_p, ct.c_char_p, ct.c_char_p, ct.c_int, ct.c_int, ct.c_char_p]
       self.token_code_lib.set_token_map.argtypes = [ct.c_char_p, ct.c_char_p, ct.c_char_p, ct.c_char_p, ct.c_char_p]
       self.token_code_lib.get_token_map.argtypes = [ct.c_char_p, ct.c_char_p, ct.c_char_p, ct.c_char_p, ct.c_char_p]
       self.token_code_lib.get_token_pad_info.argtypes = [ct.c_char_p, ct.c_char_p, ct.c_char_p, ct.c_int, ct.c_char_p, ct.c_char_p]
       self.token_code_lib.get_token_info.argtypes = [ct.c_char_p, ct.c_char_p, ct.c_char_p, ct.c_char_p, ct.c_char_p, ct.c_char_p]
       self.token_code_lib.get_reg_info.argtypes = [ct.c_char_p, ct.c_char_p, ct.c_char_p, ct.c_char_p, ct.c_char_p]
       self.token_code_lib.get_token_name.argtypes = [ct.c_char_p, ct.c_char_p, ct.c_int, ct.c_char_p, ct.c_char_p]
       self.token_code_lib.get_formula.argtypes = [ct.c_char_p, ct.c_char_p, ct.c_char_p, ct.c_char_p]
       self.token_code_lib.get_inv_formula.argtypes = [ct.c_char_p, ct.c_char_p, ct.c_char_p, ct.c_char_p]

       if verbose_messages != None:
          print_read_messages = 'yes'
       else:
          print_read_messages = 'no'
       if filename != None:
          self.load_file(filename,print_read_messages)
       else:
          self.token_code_lib.initialize('none'.encode('utf-8'))

    def __repr__(self):
       return "Nothing to report..."

    def add_token(self,reg_field_name, reg_type, token_name, bit_pattern, num_bits, default_str, description):
       # Note that reg_type is either 'register' or 'test_block'
       error_flag = self.token_code_lib.add_token(reg_field_name.encode('utf-8'),reg_type.encode('utf-8'),token_name.encode('utf-8'), bit_pattern.encode('utf-8'), ct.c_int(num_bits), default_str.encode('utf-8'), description.encode('utf-8'), self.err_msg)
       if error_flag == 1:
           print(self.err_msg.value.decode('utf-8'))
           sys.exit()
    def add_token_with_test_pad_info(self,reg_field_name, reg_type, token_name, bit_pattern, num_bits, pad0_in, pad1_in, pad2_in, pad3_in):
       pad0 = pad0_in.lower()
       pad1 = pad1_in.lower()
       pad2 = pad2_in.lower()       
       pad3 = pad3_in.lower()
       default_str = "NULL"
       description = "NULL"
       # Note that reg_type is either 'register' or 'test_block'
       error_flag = self.token_code_lib.add_token_with_test_pad_info(reg_field_name.encode('utf-8'),reg_type.encode('utf-8'),token_name.encode('utf-8'), bit_pattern.encode('utf-8'), ct.c_int(num_bits), pad0.encode('utf-8'), pad1.encode('utf-8'), pad2.encode('utf-8'), pad3.encode('utf-8'), default_str.encode('utf-8'), description.encode('utf-8'), self.err_msg)
       if error_flag == 1:
           print(self.err_msg.value.decode('utf-8'))
           sys.exit()
    def add_formula(self,reg_field_name, reg_type, formula_in, num_bits, signed_val, default_str, description):
       formula = formula_in.lower()
       # Note that reg_type is either 'register' or 'test_block'
       error_flag = self.token_code_lib.add_formula(reg_field_name.encode('utf-8'),reg_type.encode('utf-8'),formula.encode('utf-8'), num_bits, signed_val, default_str.encode('utf-8'), description.encode('utf-8'), self.err_msg)
       if error_flag == 1:
           print(self.err_msg.value.decode('utf-8'))
           sys.exit()
    def add_inv_formula(self,reg_field_name, reg_type, formula_in, num_bits, signed_val):
       formula = formula_in.lower()
       error_flag = self.token_code_lib.add_inv_formula(reg_field_name.encode('utf-8'),reg_type.encode('utf-8'),formula.encode('utf-8'), num_bits, signed_val,self.err_msg)
       if error_flag == 1:
           print(self.err_msg.value.decode('utf-8'))
           sys.exit()
    def set_token_map(self,reg_field_name, reg_type, token_name):
       # Note that reg_type is either 'register' or 'test_block'
       error_flag = self.token_code_lib.set_token_map(reg_field_name.encode('utf-8'),reg_type.encode('utf-8'),token_name.encode('utf-8'), self.out_string, self.err_msg)
       if error_flag == 1:
           print(self.err_msg.value.decode('utf-8'))
           sys.exit()
       elif error_flag == 2:
           print(self.err_msg.value.decode('utf-8'))
       out_string = self.out_string.value.decode('utf-8')
       return out_string
    def get_token_map(self,reg_field_name, reg_type, value_in):
       # Note that reg_type is either 'register' or 'test_block'
       error_flag = self.token_code_lib.get_token_map(reg_field_name.encode('utf-8'),reg_type.encode('utf-8'),value_in.encode('utf-8'), self.out_string, self.err_msg)
       if error_flag == 1:
           print(self.err_msg.value.decode('utf-8'))
           sys.exit()
       out_string = self.out_string.value.decode('utf-8')
       return out_string
    def get_pad_info(self,test_block_name, token_name, pad_num_in):
       # Note that reg_type is either 'register' or 'test_block'
       error_flag = self.token_code_lib.get_token_pad_info(test_block_name.encode('utf-8'),'test_block'.encode('utf-8'),token_name.encode('utf-8'), pad_num_in, self.out_string, self.err_msg)
       if error_flag == 1:
           print(self.err_msg.value.decode('utf-8'))
           sys.exit()
       elif error_flag == 2:
           print(self.err_msg.value.decode('utf-8'))
       out_string = self.out_string.value.decode('utf-8')
       return out_string
    def get_token_info(self,reg_name, token_name, info_str):
       # Note that reg_type is either 'register' or 'test_block'
       error_flag = self.token_code_lib.get_token_info(reg_name.encode('utf-8'),'register'.encode('utf-8'),token_name.encode('utf-8'), info_str.encode('utf-8'), self.out_string, self.err_msg)
       if error_flag == 1:
           print(self.err_msg.value.decode('utf-8'))
           sys.exit()
       elif error_flag == 2:
           print(self.err_msg.value.decode('utf-8'))
       out_string = self.out_string.value.decode('utf-8')
       return out_string
    def get_reg_info(self,reg_name, info_str):
       # Note that reg_type is either 'register' or 'test_block'
       error_flag = self.token_code_lib.get_reg_info(reg_name.encode('utf-8'),'register'.encode('utf-8'), info_str.encode('utf-8'), self.out_string, self.err_msg)
       if error_flag == 1:
           print(self.err_msg.value.decode('utf-8'))
           sys.exit()
       elif error_flag == 2:
           print(self.err_msg.value.decode('utf-8'))
       out_string = self.out_string.value.decode('utf-8')
       return out_string
    def get_reg_names(self, reg_type):
       reg_num = 0;
       out_list = []
       while(1):
          # Note that reg_type is either 'register' or 'test_block'
          error_flag = self.token_code_lib.get_reg_name(reg_type.encode('utf-8'),reg_num, self.out_string, self.err_msg)
          if error_flag == 0:
             out_list.append(self.out_string.value.decode('utf-8'))
          elif error_flag == 1:
             print(self.err_msg.value.decode('utf-8'))
             sys.exit()
          elif error_flag == 2:
             break;
          reg_num = reg_num + 1
       return out_list
    def find_reg_name(self,reg_field_search_name,reg_type):
       reg_name_list = self.get_reg_names(reg_type)
       out_list = []
       for s in reg_name_list: 
           if reg_field_search_name.lower() in s.lower():
              out_list.append(str(s))
       return(out_list)
    def print_tokens_or_formula_raw(self,reg_field_search_name,reg_type):
       token_list = self.get_token_names(reg_field_search_name,reg_type)
       if len(token_list) > 0:
           print("-> tokens are:")
           for token in token_list:
               if reg_type == 'test_block':
                   pad_mode0 = self.get_pad_info(reg_field_search_name, token, 0)
                   pad_mode1 = self.get_pad_info(reg_field_search_name, token, 1)
                   pad_mode2 = self.get_pad_info(reg_field_search_name, token, 2)
                   pad_mode3 = self.get_pad_info(reg_field_search_name, token, 3)
                   print("   pads: %s%s%s%s    token_name: %s" % \
                         (pad_mode0, pad_mode1, pad_mode2, pad_mode3, token)) 
               else:
                 print("   %s" % token)
       else:
           formula_list = self.get_formula(reg_field_search_name,reg_type)
           if len(formula_list) > 0:
               print("-> no tokens, but formula is '%s'" % formula_list[0])
               inv_formula_list = self.get_inv_formula(reg_field_search_name,reg_type)
               if len(inv_formula_list) > 0:
                   out_min = self.get_token_map(reg_field_search_name,reg_type,'min')
                   out_max = self.get_token_map(reg_field_search_name,reg_type,'max')
                   print("          and inv_formula is '%s'" % inv_formula_list[0])
                   print("          Note: min formula input is '%s'" % out_min)
                   print("                max formula input is '%s'" % out_max)
               else:
                   print("  Warning: no inv_formula exists")
           else:
               inv_formula_list = self.get_inv_formula(reg_field_search_name,reg_type)
               if len(inv_formula_list) > 0:
                   print("-> Warning: no tokens and no formula exists, but inv_formula is '%s'" % inv_formula_list[0])
               else:
                   print("-> no tokens, formula, or inv_formula exist")
       
    def print_reg_tokens(self,reg_field_search_name):
       reg_name_list = self.find_reg_name(reg_field_search_name,'register')
       for s in reg_name_list:
           if s == reg_field_search_name:
               reg_name_list = [s]
       if len(reg_name_list) == 0:
           print("\nNo register name matches '%s'" % reg_field_search_name)
       elif len(reg_name_list) > 1:
           print("\nMore than one register name matches '%s':" % reg_field_search_name)
           for s in reg_name_list:
               print("   %s" % s)
       else:
           print("\nRegister name '%s' is a match to '%s'" % (reg_name_list[0],reg_field_search_name))
           self.print_tokens_or_formula_raw(reg_name_list[0],'register')

    def print_test_tokens(self,test_block_search_name):
       test_block_name_list = self.find_reg_name(test_block_search_name,'test_block')
       for s in test_block_name_list:
           if s == test_block_search_name:
               test_block_name_list = [s]

       if len(test_block_name_list) == 0:
           print("\nNo test block name matches '%s'" % test_block_search_name)
       elif len(test_block_name_list) > 1:
           print("\nMore than one test_block name matches '%s':" % test_block_search_name)
           for s in test_block_name_list:
               print("   %s" % s)
       else:
           print("\nTest block name '%s' is a match to '%s'" % (test_block_name_list[0],test_block_search_name))
           self.print_tokens_or_formula_raw(test_block_name_list[0],'test_block')

    def print_reg(self,reg_field_search_name):
       reg_name_list = self.find_reg_name(reg_field_search_name,'register')
       if len(reg_name_list) == 0:
           print("\nNo register name matches '%s'" % reg_field_search_name)
       elif len(reg_name_list) > 1:
           print("\nMore than one register name matches '%s':" % reg_field_search_name)
           for s in reg_name_list:
               print("   %s" % s)
       else:
           print("\nRegister name '%s' is a match to '%s'" % (reg_name_list[0],reg_field_search_name))

    def print_test(self,test_block_search_name):
       test_block_name_list = self.find_reg_name(test_block_search_name,'test_block')
       if len(test_block_name_list) == 0:
           print("\nNo test block name matches '%s'" % test_block_search_name)
       elif len(test_block_name_list) > 1:
           print("\nMore than one test block name matches '%s':" % test_block_search_name)
           for s in test_block_name_list:
               print("   %s" % s)
       else:
           print("\nTest block name '%s' is a match to '%s'" % (test_block_name_list[0],test_block_search_name))
    def get_token_names(self,reg_field_name,reg_type):
       token_num = 0;
       out_list = []
       while(1):
          # Note that reg_type is either 'register' or 'test_block'
          error_flag = self.token_code_lib.get_token_name(reg_field_name.encode('utf-8'),reg_type.encode('utf-8'),token_num, self.out_string, self.err_msg)
          if error_flag == 0:
             out_list.append(self.out_string.value.decode('utf-8'))
          elif error_flag == 1:
             print(self.err_msg.value.decode('utf-8'))
             sys.exit()
          elif error_flag == 2:
             break;
          token_num = token_num + 1
       return out_list
    def get_formula(self,reg_field_name, reg_type):
       out_list = []
       # Note that reg_type is either 'register' or 'test_block'
       error_flag = self.token_code_lib.get_formula(reg_field_name.encode('utf-8'),reg_type.encode('utf-8'), self.out_string, self.err_msg)
       if error_flag == 0:
           out_list.append(self.out_string.value.decode('utf-8'))
       elif error_flag == 1:
           print(self.err_msg.value.decode('utf-8'))
           sys.exit()
       return out_list
    def get_inv_formula(self,reg_field_name, reg_type):
       out_list = []
       # Note that reg_type is either 'register' or 'test_block'
       error_flag = self.token_code_lib.get_inv_formula(reg_field_name.encode('utf-8'),reg_type.encode('utf-8'), self.out_string, self.err_msg)
       if error_flag == 0:
           out_list.append(self.out_string.value.decode('utf-8'))
       elif error_flag == 1:
           print(self.err_msg.value.decode('utf-8'))
           sys.exit()
       return out_list
    def print_registers(self):
       self.token_code_lib.print_registers()
    def load_file(self,filename,print_read_messages):
        if os.path.isfile(filename) == False:
            print("Error using loadfile:  file '%s' does not exist\n" % filename)
            sys.exit()
        self.token_code_lib.initialize(filename.encode('utf-8'))

        wb = openpyxl.load_workbook(filename)
        sheet = wb.get_sheet_by_name('RegTokens')
        if print_read_messages == 'yes':
            print("\n... Read in RegTokens sheet ...")
        for row in range(2, sheet.max_row + 1):
            skip_flag = 0
            reg_field_name_raw = sheet['A' + str(row)].value
            if reg_field_name_raw is None:
               skip_flag = 1
            else:
               reg_field_name = str(reg_field_name_raw).strip('" ')
            num_bits_raw = sheet['B' + str(row)].value
            if num_bits_raw is None:
               skip_flag = 1
            else:
               num_bits = int(num_bits_raw)
            entry_type_raw = sheet['C' + str(row)].value
            if entry_type_raw is None:
               skip_flag = 1
            else:
               entry_type = str(entry_type_raw).strip('" ')
            token_val_raw = sheet['D' + str(row)].value
            if token_val_raw is None:
               skip_flag = 1
            else:
               token_val = str(token_val_raw).strip('" ')
               formula_expr = token_val
            bit_seq_raw = sheet['E' + str(row)].value
            if bit_seq_raw is None:
               skip_flag = 1
            else:
               bit_seq = str(bit_seq_raw).strip('" ')
               signed_val = 1 if bit_seq == 'signed' else 0
            default_str_raw = sheet['F' + str(row)].value
            if default_str_raw is None:
               default_str = 'NULL'
            else:
               default_str = str(default_str_raw).strip('" ')
            description_raw = sheet['G' + str(row)].value
            if description_raw is None:
               description = 'NULL'
            else:
               description = str(description_raw).strip('" ')

#            print("reg = %s, num_bits = %d, entry = %s, token/formula = %s, bit_seq = %s, signed_val = %d"%(reg_field_name, num_bits, entry_type, token_val, bit_seq, signed_val))
            if skip_flag == 1:
                if print_read_messages == 'yes':
                   print("   -> skipping row %d in RegTokens sheet" % row)
            else:
                if entry_type == 'token':
                   self.add_token(reg_field_name,'register',token_val,bit_seq,num_bits,default_str,description)
                elif entry_type == 'formula':
                   self.add_formula(reg_field_name,'register',formula_expr,num_bits,signed_val,default_str,description)
                elif entry_type == 'inv_formula':
                   self.add_inv_formula(reg_field_name,'register',formula_expr,num_bits,signed_val)

        sheet = wb.get_sheet_by_name('TestTokens')
        if print_read_messages == 'yes':
            print("\n... Read in TestTokens sheet ...")
        for row in range(2, sheet.max_row + 1):
            skip_flag = 0
            reg_field_name_raw = sheet['A' + str(row)].value
            if reg_field_name_raw is None:
               skip_flag = 1
            else:
               reg_field_name = str(reg_field_name_raw).strip('" ')
            num_bits_raw = sheet['B' + str(row)].value
            if num_bits_raw is None:
               skip_flag = 1
            else:
               num_bits = int(num_bits_raw)
            entry_type_raw = sheet['C' + str(row)].value
            if entry_type_raw is None:
               skip_flag = 1
            else:
               entry_type = str(entry_type_raw).strip('" ')
            token_val_raw = sheet['D' + str(row)].value
            if token_val_raw is None:
               skip_flag = 1
            else:
               token_val = str(token_val_raw).strip('" ')
            bit_seq_raw = sheet['E' + str(row)].value
            if bit_seq_raw is None:
               skip_flag = 1
            else:
               bit_seq = str(bit_seq_raw).strip('" ')
            pad0_raw = sheet['H' + str(row)].value
            if pad0_raw is None:
               pad0 = "f"
            else:
               pad0 = str(pad0_raw).strip('" ')
            pad1_raw = sheet['I' + str(row)].value
            if pad1_raw is None:
               pad1 = "f"
            else:
               pad1 = str(pad1_raw).strip('" ')
            pad2_raw = sheet['J' + str(row)].value
            if pad2_raw is None:
               pad2 = "f"
            else:
               pad2 = str(pad2_raw).strip('" ')
            pad3_raw = sheet['K' + str(row)].value
            if pad3_raw is None:
               pad3 = "f"
            else:
               pad3 = str(pad3_raw).strip('" ')
#            print("reg = %s, num_bits = %d, entry = %s, token = %s, bit_seq = %s"%(reg_field_name, num_bits, entry_type, token_val, bit_seq))
            if skip_flag == 1:
                if print_read_messages == 'yes':
                    print("   -> skipping row %d in TestTokens sheet" % row)
            else:
                if entry_type == 'token':
                    self.add_token_with_test_pad_info(reg_field_name,'test_block',token_val,bit_seq,num_bits,pad0,pad1,pad2,pad3)
